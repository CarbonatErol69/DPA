#TODO: implement connection to database

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl
from openpyxl import load_workbook
import pandas as pd

class EventSchedulerApp(tk.Tk):
#UI options
    
    def __init__(self):
        super().__init__()
        self.title("Super tolle App! Mit ohne Funktionalität!")
        self.geometry("500x500")
        #TODO: fix the window not being dragable after windows titel removal
        #self.overrideredirect(True) #removes windows title bar, but cant be dragged afterwards
        self.resizable(1,1)

        # Initialize Excel workbook
        self.excel_file = r"C:\Users\DE88342\OneDrive - Grunenthal Group\Desktop\VSC - Py\DPAvsBWV\schedulize_db.xlsx"
        self.workbook = openpyxl.load_workbook(self.excel_file)
        self.sheet = self.workbook["Student"]

        # Initialize variables to store data
        self.students = []
        self.events = []

        # Create and place widgets
        self.create_widgets()

    def create_widgets(self):
        entry_font = {'font': ('Helvetica', 11)}
        paddings = {'padx': 5, 'pady': 5}

        #enter surname
        self.label_student_name = tk.Label(self, text="Name:")
        self.entry_student_name = tk.Entry(self, **entry_font)
        self.label_student_name.grid(row=0, column=0, sticky="w", **paddings)
        self.entry_student_name.grid(row=0, column=1)

        #enter first name
        self.label_first_name = tk.Label(self, text="Vorname:")
        self.entry_first_name = tk.Entry(self, **entry_font)
        self.label_first_name.grid(row=1, column=0, sticky="w", **paddings)
        self.entry_first_name.grid(row=1, column=1)

        #enter klasse
        self.label_class = tk.Label(self, text="Klasse:")
        self.entry_class = tk.Entry(self, **entry_font)
        self.label_class.grid(row=2, column=0, sticky="w", **paddings)
        self.entry_class.grid(row=2, column=1)

        #enter preferences
        self.label_event_preference = tk.Label(self, text="Event Wahl (bis zu 6):")
        
        #this is the hard way to do whats shorter afterwards
        '''self.entry_event_preference1 = tk.Entry(self)
        self.entry_event_preference2 = tk.Entry(self)
        self.entry_event_preference3 = tk.Entry(self)
        self.entry_event_preference4 = tk.Entry(self)
        self.entry_event_preference5 = tk.Entry(self)
        self.entry_event_preference6 = tk.Entry(self)'''
        self.entry_event_preferences = [tk.Entry(self, **entry_font) for _ in range(6)]
        for i, entry in enumerate(self.entry_event_preferences):
            entry.grid(row= i+4, column = 1, **paddings) 
            #i+4 statt i+1 da wir grid layout manager benutzen
        #TODO: Ohrfeige, 30 min gebraucht um columns zu column zu machen um bug zu fixen

        self.label_event_preference.grid(row=3, column=0, sticky="w", **paddings)
        #self.entry_event_preference1.grid(row=1, column=1)
        #self.entry_event_preference2.grid(row=2, column=1)
        #self.entry_event_preference3.grid(row=3, column=1)
        #self.entry_event_preference4.grid(row=4, column=1)
        #self.entry_event_preference5.grid(row=5, column=1)
        #self.entry_event_preference6.grid(row=6, column=1)

        self.button_add_student = tk.Button(self, text="hIEr KlICkEN füR vIRusSCan FREE", command=self.add_student)
        self.button_add_student.grid(row=10, column=1, columnspan=2, **paddings)

    def add_student(self):
        # Retrieve student information from entry fields
        student_name = self.entry_student_name.get()
        event_preferences = [entry.get() for entry in self.entry_event_preferences]
        

        #code to change, incase get.event_prefs doesnt work
        '''[
            self.entry_event_preference1.get(),
            self.entry_event_preference2.get(),
            self.entry_event_preference3.get(),
            self.entry_event_preference4.get(),
            self.entry_event_preference5.get(),
            self.entry_event_preference6.get(),
        ]'''

        # Append student to list
        self.students.append((student_name, event_preferences))

        # Write student data to Excel file
        self.write_student_to_excel(student_name, event_preferences)

        # Clear entry fields
        self.entry_student_name.delete(0, tk.END)
        for entry in self.entry_event_preferences:
            entry.delete(0, tk.END)
            
        #clear entry fields (complete, unshortened)
        #self.entry_student_name.delete(0, tk.END)
        #self.entry_event_preference1.delete(0, tk.END)
        #self.entry_event_preference2.delete(0, tk.END)
        #self.entry_event_preference3.delete(0, tk.END)
        #self.entry_event_preference4.delete(0, tk.END)
        #self.entry_event_preference5.delete(0, tk.END)
        #self.entry_event_preference6.delete(0, tk.END)

    '''      print("Writing student data to Excel...")
        print("Name:", name)
        print("Preferences:", preferences)
        next_row = self.sheet.max_row + 1
        print("Next Row:", next_row)

        self.sheet.cell(row=next_row, column=1, value=name)
        for col, pref in enumerate(preferences, start=2):
            self.sheet.cell(row=next_row, column=col, value=pref)
        self.workbook.save(self.excel_file)
        print("Data written to Excel successfully.")'''

    #test, to determine where data is lost during writing (to xslx) process
        #TODO: Clear this method, when its not needed anymore
    def write_student_to_excel(self, name, preferences):
    # Append student data to Students sheet in Excel file
        df = pd.DataFrame({'Student Name': [name], 'Event Preference 1': preferences[0],
                           'Event Preference 2': preferences[1], 'Event Preference 3': preferences[2],
                           'Event Preference 4': preferences[3], 'Event Preference 5': preferences[4],
                           'Event Preference 6': preferences[5]})

        # Append DataFrame to Excel file
        #with pd.ExcelWriter(self.excel_file, mode='a', engine='openpyxl') as writer:
        writer = pd.ExcelWriter(self.excel_file, mode = 'a', engine = 'openpyxl')
            #load the existing workbook
        writer.book = load_workbook(self.excel_file)
        sheet_name = 'Student'

        if sheet_name in writer.book.sheetnames:
            sheet = writer.book[sheet_name]
                #find the last row with data
            start_row = sheet.max_row + 1
        else:
                #if the sheet doesnt exist then create it
            start_row = 0 #start from the beginning

             #write the new data to the sheet   
        df.to_excel(writer, sheet_name='Student', index=False, header=True, startrow=start_row)
        #save the workbook
        writer.book.save(self.excel_file)

        #close the writer to release resources
        writer.close()

    def read_students_from_excel(self):
        # Read student data from Student sheet in Excel file
        sheet = self.workbook["Student"]
        for row in sheet.iter_rows(values_only=True):
            student_name = row[0]
            event_preferences = row[1:]
            self.students.append((student_name, event_preferences))

    # Add methods for event scheduling, display results, etc.
    def schedule_events(self):
        # Implement event scheduling logic here
        pass

    def display_timetable(self):
        # Implement timetable display logic here
        pass

    def display_attendance_list(self):
        # Implement attendance list display logic here
        pass

if __name__ == "__main__":
    app = EventSchedulerApp()
    app.mainloop()
